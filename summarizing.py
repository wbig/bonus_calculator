import pandas as pd
import os
import re
import warnings
import calendar
import numpy
import shutil
from collections import Counter
from tabulate import tabulate
from pathlib import Path
from datetime import datetime
from docx import Document
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, GradientFill, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# 忽略运行时警告
warnings.filterwarnings("ignore", category=RuntimeWarning)
# 忽略 FutureWarning
warnings.filterwarnings("ignore", category=FutureWarning)

# 机器人8%全给护士，PRP算绩效后医生平分，平日挂号收入和会诊收入给医生


def ctrl_panel():

    """获取路径"""
    path = set_path()

    while True:
        choices = """
    ==================================================================================================================
    请选择要进行的操作：\n
        1. 患者归属检查    2. 收入数据汇总    3. 罚款数据查询    4. 整合收入和罚款    5. 奖金计算    6. 查看总值班人员
        0. 重设路径        Q. 退出
    ==================================================================================================================
        """

        choice = input(choices)

        # 创建一个字典来映射选择和对应的函数
        operations = {
            '1': patient_affiliation_check,
            '2': output_revenue_data,
            '3': query_fine_data,
            '4': Data_sum,
            '5': bonus_calculation,
            '6': extract_namelist_from_word,
            '0': set_path
        }

        if choice.upper() == 'Q':
            print("退出程序...")
            break  # 退出循环

        # 检查用户输入是否在字典的键中
        if choice in operations:
            # 调用对应的函数
            if choice == '0':
                path = operations[choice](choice)
            else:
                operations[choice](path)
        else:
            print("无效的选择，请重新输入！")


# 确认数据文件夹内存在 groups.xlsx 文件，没有则复制
def groups_path(script_dir, path_data):

    path_destination = os.path.join(path_data, "groups.xlsx")
    # 检查目标文件是否已经存在
    if not os.path.exists(path_destination):
        # 获取groups文件的绝对路径
        path_groups_src = os.path.join(script_dir, r"lists\groups.xlsx")
        shutil.copy(path_groups_src, path_destination)
        print('\nThe file "groups.xlsx" is copied to the data directory.')

    return path_destination

    
def get_df_groups(path):
    
    """获取脚本所在的目录"""
    script_dir = get_abspath()
    path_groups = groups_path(script_dir, path)
    check_if_exists(path_groups)
    df_groups = pd.read_excel(path_groups)
    
    return df_groups

"""全科成员（包括医生及护士）信息在此，如有成员调整，需要要修改 groups.xlsx """
def staff_info(path):

    df_groups = get_df_groups(path)

    df_doctors_info = df_groups[["G", "D", "S", "PG", "PW"]].\
        dropna(subset="D").fillna(1)
    df_nurses_info = df_groups[["N"]].dropna()

    # 筛选出D列名字长度小于等于3的行
    """所有医生"""
    # 检查 D 列元素是否为非空(大于1)字符串并且长度不超过3
    df_doctors_info = df_doctors_info[df_doctors_info['D'].apply(
        lambda x: isinstance(x, str) and len(x) > 1 and len(x) <= 3)]

    # 检查是否有重复行
    duplicates = df_doctors_info[df_doctors_info.duplicated(keep=False)]  # 找出所有重复行

    if not duplicates.empty:
        print("发现重复行，将删除重复行！")
        print("重复行如下：")
        print(duplicates)
        df_doctors_info = df_doctors_info.drop_duplicates()  # 删除重复行，保留第一个副本

    # 检查D列是否有重复值
    if df_doctors_info['D'].duplicated().any():
        print("\n警告：D列中有重复值，请检查！")
        return

    # 找到 S 列中值为 "均奖" 的行
    mask1 = df_doctors_info['S'] == '均奖'
    # 将这些行的 G 列的值改为 "avg"
    df_doctors_info.loc[mask1, 'G'] = 'avg'

    # 找到 PW 列中值为 0 的行
    mask1 = (df_doctors_info['PW'] < 1) & (df_doctors_info['PW'] > 0)
    df_doctors_info.loc[mask1, 'G'] = 'partial'

    # 找到 零奖者 的行
    mask2 = (df_doctors_info['S'] == '不在岗')\
        | (df_doctors_info['PG'] == 0)\
        | (df_doctors_info['PW'] == 0)

    df_doctors_info.loc[mask2, 'G'] = 'zero'
    df_doctors_info.loc[mask2, 'S'] = '不在岗'
    
    # 分别返回医生和护士成员信息
    return df_nurses_info, df_doctors_info


# 使用groupby对G列进行分组，并对D列使用list函数进行聚合
# 对D列和特殊人群（special）分组
# 根据不同的工作或奖金分配状态分组

def info_df_gb(df_doctors_info):

    def groupby_group_named(df, group_col, agg_dict):
        if df.empty:
            print("The DataFrame is empty.")
            return df

        df_out = df.groupby(group_col).agg(**agg_dict)
        return df_out

    # 定义命名聚合字典
    agg_dict = {
        'D_list': ('D', list),
        'D_count': ('D', 'count'),
        'PW': ('PW', 'sum'),
        'PG': ('PG', 'sum')
    }

    # 在岗者(包括返聘)
    df_在岗 = df_doctors_info[df_doctors_info.G != "zero"]
    gb_在岗 = groupby_group_named(df_在岗, "G", agg_dict)

    # 在勤者(不包括返聘)
    df_在勤 = df_在岗[df_在岗.S != "返聘"]
    gb_在勤 = groupby_group_named(df_在勤, "G", agg_dict)

    # 夜班者
    agg_dict_2 = {
        'D_list': ('D', list),
        'D_count': ('D', 'count')
    }
    staff_lst = df_doctors_info.D.tolist()
    df_夜班 = df_在勤[(df_在勤.S == 1) & df_在勤.G.isin(staff_lst)]
    gb_夜班 = groupby_group_named(df_夜班, "G", agg_dict_2)

    return df_在岗, gb_在岗, df_在勤, gb_在勤, df_夜班, gb_夜班


# Obtain the path to the data file
def set_path(choice=None):

    pathFile = get_abspath("path.txt")

    # 确认路径存在且格式正确（包含年月信息且以其结尾）
    def confirm_path(path):
        while True:
            if not os.path.exists(path):
                print(f'The directory does not exist. Please reset.\n')
                path = input('Please enter new directory ("Q" for return):')
                if path.upper() == "Q":
                    return path_original
            else:
                # 使用正则表达式匹配年份和月份，确认文件夹格式正确
                pattern = r"^\d{4}年\d{1,2}月$"
                match = re.search(pattern, os.path.basename(path))

                if match:
                    with open(pathFile, "w", encoding="UTF8") as f:
                        f.write(path)
                    print("Path setting succeeded.")
                    break
                else:
                    path = input('The path format is incorrect, please re-enter ("Q" for return):')
                    if path.upper() == "Q":
                        return path_original
        return path

    if os.path.exists(pathFile):
        with open(pathFile, "r", encoding="UTF8") as f:
            path_original = f.read().strip()
    else:
        path_original = "None"

    # renew the derictory
    if choice == "0":
        path_new = input('Please enter new directory ("Q" for return):')
        if path_new.upper() != "Q":
            path_new = confirm_path(path_new)
            print(f'\nCurrent Directory: {path_new}')
            return path_new
        else:
            print(f'\nCurrent Directory: {path_original}')
            return path_original

    print(f'\nCurrent Directory: {path_original}')
    path = confirm_path(path_original)

    return path


# Gets the absolute path to the derictory or file
def get_abspath(relative_path=None):

    # 获取当前脚本文件的绝对路径
    script_path = os.path.abspath(__file__)
    # 获取脚本所在的目录
    script_dir = os.path.dirname(script_path)

    if relative_path is not None:
        # 获取目标文件的绝对路径
        abspath = os.path.join(script_dir, relative_path)
        return abspath

    return script_dir


# Check whether the file exists
def check_if_exists(path_file):

    while not os.path.exists(path_file):
        print("\nFile not found at path: {}".format(path_file))
        input("Please check the path and try again.\n")


# patient affiliation check
def patient_affiliation_check(path):

    df_groups = get_df_groups(path)[["G", "D"]].dropna(subset="D")

    # 获取医生列表
    list_doctor = df_groups.D.unique().tolist()

    path_data = os.path.join(path, r"ZLHIS\吉林省人民医院病人记帐费用汇总清单.xlsx")
    check_if_exists(path_data)
    wb = load_workbook(path_data)
    ws = wb.active

    cell_date = ws.cell(2, 1).value
    pattern = r"\d{4}.*至.*日"
    range_date = re.search(pattern, cell_date).group()

    r = ws.max_row
    # c = ws.max_column

    # 获取表格的有效长度 row
    for row in range(1, r):
        data = ws.cell(row, 1).value
        if (type(data) is str) and ("合计" in data):
            row = row
            break

    rn = 0
    rc = 0
    fill_0 = PatternFill(fill_type='solid', fgColor='B0F8B0')
    fill_1 = PatternFill(fill_type='solid', fgColor='FFFF00')

    # 正则表达式模式，去除空格或符合模式的工号
    def replace_employee_id(pattern, cell):

        # 检查单元格值是否不是 None
        if cell.value is not None:
            # 将单元格值转换为字符串
            cell_value_str = str(cell.value)
            # 使用正则表达式去除单元格的 pattern 内容
            replaced_cell = re.sub(pattern, '', cell_value_str)
        else:
            replaced_cell = ""

        return replaced_cell

    for row in range(4, row):

        pattern = r'\s+|\w\d{4}'

        r_d = ws.cell(row, 5)
        # 住院医生姓名
        resident_doctor = replace_employee_id(pattern, r_d)

        o_p = ws.cell(row, 4)
        # 门诊医生姓名
        outpatient_physician = replace_employee_id(pattern, o_p)

        # 替换医生信息
        if resident_doctor not in list_doctor:
            if outpatient_physician in list_doctor:
                # 单元格赋新值并将原住院医师信息插入批注
                comment = Comment(resident_doctor, "原始信息：")
                ws.cell(row, 5, value=outpatient_physician).comment = comment
                r_d.fill = fill_0
                rn += 1
            else:
                r_d.fill = fill_1
                rc += 1
    destination_path = r"ZLHIS\吉林省人民医院病人记帐费用汇总清单_ForUse.xlsx"
    destination_path = os.path.join(path, destination_path)
    print("\n吉林省人民医院病人记帐费用汇总清单_ForUse.xlsx ... ", end="")
    wb_save(wb, destination_path)
    wb.close()

    os.system(f'start excel "{destination_path}"')

    print(f"\n日期范围：{range_date}")
    print("\n当前分组情况：", df_groups.groupby("G")["D"].agg(list), sep="\n")
    print(f"\n归属检查完毕：更正了{rn}条信息，有{rc}个需要进一步核实。请验证!")


# Get the list of chief duty personnel and the amount of doctor duty fee
def extract_namelist_from_word(path):

    df_groups = get_df_groups(path)

    doctors = df_groups.D.unique().tolist()
    nurses = df_groups.N.unique().tolist()

    # 总值班人员 = 全科人员 - 排除列表
    排除列表 = ["于海涛", "刘颖"]  # 院内有重名，故需排除

    # 定义输出路径
    path_txt_out = '总值名单.txt'
    outpu_path = os.path.join(path, path_txt_out)

    if os.path.exists(outpu_path):
        with open(outpu_path, "r", encoding="utf-8") as f:
            content = f.read()
            if content:
                创二总值 = content.split(",")
                print(f'\n"{path_txt_out}" 文件已存在。')

    else:

        list_duty_mamager = [name for name in doctors + nurses if name not in 排除列表]

        pattern = r"(\d{4})年(\d{1,2})月"
        match = re.search(pattern, path)

        year = match[1]
        month = int(match[2])

        name_file = f"{year}年{month}月份值班表.docx"
        path_docx_in = os.path.join(path, name_file)
        check_if_exists(path_docx_in)

        doc = Document(path_docx_in)

        # 提取人名
        创二总值 = []
        for table in doc.tables:
            for row in table.rows:  # 使用enumerate()函数获取行索引
                for cell in row.cells:  # 使用enumerate()函数获取列索引
                    name = cell.text.replace(" ", "")
                    if name in list_duty_mamager:
                        创二总值.append(name)

        with open(outpu_path, "w", encoding='utf8') as f:
            text = ','.join(创二总值)
            f.write(text)
            print(f'\n"{path_txt_out}" 文件已创建。')

    # income_penalty.xlsx 中填入总值金额
    path_income_penalty = os.path.join(path, "income_penalty.xlsx")
    check_if_exists(path_income_penalty)

    df_income_penalty = pd.read_excel(path_income_penalty,
                                      index_col="收入和处罚").fillna(0)
    医生总值 = [i for i in 创二总值 if i not in nurses]
    # 填入总值金额
    df_income_penalty.loc["总值", "金额"] = len(医生总值) * 50

    医生总值班费 = len(医生总值) * 50

    print(f"创二总值：{创二总值}")

    return 医生总值班费


# Get the PRP amount
def getPRP(path):

    pathPRP = os.path.join(path, r"ZLHIS\创伤骨病二科门诊.xlsx")
    check_if_exists(pathPRP)
    df = pd.read_excel(pathPRP, skiprows=3)
    df["收入项目"] = df["收入项目"].str.strip()
    df = df.set_index("收入项目")
    if "合计" not in df.columns or "治疗收入" not in df.index:
        return 0
    PRP = df["合计"]["治疗收入"]

    return PRP


# Get date range
def get_Date_range(path):

    path_data = os.path.join(path, r"ZLHIS\吉林省人民医院病人记帐费用汇总清单.xlsx")
    check_if_exists(path_data)
    df = pd.read_excel(path_data)
    v = df.iloc[:2, :1].to_string(index=False)
    pattern = r"\d{4}年\d{1,2}月\d{1,2}日"
    dates = re.findall(pattern, v)

    # 将字符串转换为datetime对象
    date_obj = datetime.strptime(dates[0], "%Y年%m月%d日")

    # 提取年和月
    year = date_obj.year
    month = date_obj.month
    # 获取该月的天数
    days_in_month = calendar.monthrange(year, month)[1]

    dates = "_".join(dates)
    return dates, days_in_month


# output df_null
def df_null(msg):

    df_null = pd.DataFrame()
    df_null.loc["总计", "合计"] = 0
    df_null.loc["总计", msg] = 0
    print(f'暂无"{msg}"数据。\n')
    return df_null


# Remove the Spaces and employee numbers of all cells in df
def wipe(df):

    df = df.copy()
    # 正则表达式模式，匹配空格或符合模式的工号
    pattern = r'\s+|\w\d{4}'
    pd.set_option('future.no_silent_downcasting', True)
    df = df.replace(pattern, '', regex=True)
    return df


# Summarize df data to obtain sum and ratio
def summarize(df, msg, df_groups):

    # 先去除“住院医师”列中单元格的所有空格和工号信息
    df = wipe(df)

    # 合并住院医师到组的映射
    df = df.merge(df_groups[['D', 'G']],
                  left_on='住院医师',
                  right_on='D', how='left')

    # 去除 G 列中的空值
    df.dropna(subset=['G'], inplace=True)
    # 在 pandas 中，infer_objects() 方法用于推断 df 中每一列的数据类型，并尝试将其转换为更合适的类型。
    df = df.infer_objects(copy=True)
    df = df.fillna(0)
    # 聚合
    df = df.groupby("G").sum(numeric_only=True)
    # 按行求和，忽略NaN值
    df['合计'] = df.sum(axis=1, skipna=True)
    # 仅保留合计列
    df = df[['合计']]
    # 合计列求和
    total = df['合计'].sum()
    df.loc['总计'] = total
    df[msg] = df['合计'] / total if total != 0 else 0

    if msg == "罚款到组":
        print(f'{msg} 。。', end="")
    else:
        print(f'{msg} Done!')
    return df


# All summary results are set to background_gradient format
def set_gradient_style(df_merged):

    columns = df_merged.columns
    # 收入内容列表
    subsets = [c for c in columns if c != "合计"]
    # 修改重复列索引名“合计”,添加序号
    columns = ["合计_" + str(int(i/2) + 1) if c == "合计" else c
               for i, c in enumerate(columns)]

    df_merged.columns = pd.Index(columns)  # 传回索引对象
    df_merged.index.name = 'G'

    df_merged = df_merged.style.background_gradient(
        subset=subsets, cmap='YlGnBu')
    return df_merged


# Function DataFrame save to Excel
def save_xlsx(df, path):

    try:
        df.to_excel(path, index=True)
        print("(sv) Done.")

    except PermissionError:
        print("Error(sv): The file is currently in use by another program.",
            "Please close the file and press enter.")
        input()
        save_xlsx(df, path)  # Recursive call to retry saving the file

    except Exception as e:
        print(f"(sv) An error occurred: {e}")


# Function wb save to Excel
def wb_save(wb, path):

    try:
        wb.save(path)
        print("(wb) Done.")
    except PermissionError:
        print("Error(wb): The file is currently in use by another program.",
            "Please close the file and press enter.")
        input()
        wb_save(wb, path)  # Recursive call to retry saving the file
    except Exception as e:
        print(f"(wb) An error occurred: {e}")


# 获取三医联动和目标值内的奖罚数据
def query_fine_data(path):

    # 以下定义开头元素列表
    fine_list = ["合计", "金额", "处罚", "扣款"]
    reward_list = ["奖励"]

    df_groups = get_df_groups(path)

    # 所有医生列表
    lst_doctors = df_groups.D.unique().tolist()
    

    # 包含本科室及科室的列表
    my_dep = ["创伤骨病二科", "创伤骨二科", "创伤骨二", "创二", "骨二"]
    lst_dep = ['科室', '出院科室', '转入科室']

    # 在  filter 中，初始化筛选出的 df 的序号
    name_column = 0
    # 所有本科室相关的df
    total_df_list = []
    # 具体到医生的df
    personalized_df_list = []
    # 有具体数值的df
    total_lst = []

    def if_no_columns(df, file_name, sheet_name):
        # 定义一个function处理无columns的df，填加加一个含有科室和住院医师的columns
        # 定义一个子function查找df中是否有列表中的元素，如果有则输出列号
        def add_columns(df, list):
            y, x = df.shape
            for r in range(y):
                for c in range(x):
                    v = df.iat[r, c]
                    v = str(v).strip()
                    if v in list:
                        col = c
                        return col
        new_columns = [str(i) for i in range(df.shape[1])]
        # 查找是否有本科室的行
        col_dep = add_columns(df, my_dep)
        if col_dep is not None:
            new_columns[col_dep] = '科室'

            # 查找是否有本科室医生的行
            col_doc = add_columns(df, lst_doctors)
            if col_doc is not None:
                new_columns[col_doc] = '住院医师'

            original_columns = df.columns.tolist()
            # 更新 DataFrame 的列名
            df.columns = new_columns
            # 将原有列名作为一个新行添加到 DataFrame 的顶部
            new_row = pd.DataFrame([original_columns], columns=new_columns)
            df = pd.concat([new_row, df], ignore_index=True)
            print(file_name, sheet_name, 'No columns index, added one.')
            return df

    def df_spliter(df):
        split_dfs = []
        y, x = df.shape
        # 找出可做为columns的行，之后做为columns
        dep_contained_rows = []
        for r in range(y):
            for c in range(x):
                v = df.iloc[r, c]
                if str(v) in lst_dep:
                    dep_contained_rows.append(r)

        len_dep = len(dep_contained_rows)

        if len_dep == 1:
            split_dfs.append(df)

        elif len_dep > 1:
            # 检查索引范围并进行切片
            for i in range(len_dep - 1):
                # 对于第一个到倒数第二个科室行，切片从当前科室行到下一个科室行之间的内容
                start_index = dep_contained_rows[i]
                end_index = dep_contained_rows[i + 1]
                df_slice = df.iloc[start_index:end_index, :]
                split_dfs.append(df_slice)
            # 处理最后一个切片
            # 对于最后一个科室行，切片从这行到数据帧末尾
            last_start_index = dep_contained_rows[-1]
            df_slice = df.iloc[last_start_index:, :]
            split_dfs.append(df_slice)
        return split_dfs

    def filter(df, file_name, sheet_name, header):
        df_out = None

        def sub_filter(v):

            df[v] = df[v].ffill().str.strip()
            df_out = df[df[v].isin(my_dep)]
            return df_out

        df = df.copy()

        # 通过第一个单元格(header)筛选除去非数据sheet
        # ['Empty1stRow AddCol', '1stRowIsNotCol AddCol', '1stRowIsCol']

        header_filter_pattern = r"1stRow"
        if not re.search(header_filter_pattern, header):
            y, x = df.shape
            for r in range(y):
                for c in range(x):
                    v = df.iloc[r, c]
                    if str(v) in lst_dep:
                        # 将含有科室的行做为列索引
                        df.columns = df.iloc[r,]
                        df.columns.name = ''
                        df_out = sub_filter(v)
        else:
            for v in df.columns:
                if v in lst_dep:
                    df_out = sub_filter(v)

        if df_out is not None and len(df_out) > 0:
            # 对 DataFrame 中的所有字符串元素进行前后空格去除
            df_out = df_out.map(lambda x: x.strip() if
                                          isinstance(x, str) else x)
            # 计算每一行非缺失值的数量
            row_valid_counts = df_out.notnull().sum(axis=1)
            # 原地修改，保留非空单元格大于3个的行
            df_out = df_out[row_valid_counts >= 3]

            # print(file_name)
            personalized(df_out)

            # 当你需要在嵌套函数中修改外层函数的局部变量时，可以使用nonlocal
            nonlocal name_column
            name_column += 1
            df_out = df_out.copy()
            # 在df最前增加一列，并在第一个单元格内写入h值
            # DataFrame.insert(loc, column, value, allow_duplicates=False)
            df_out.insert(0, name_column, None)
            tail_line_number = df_out.index[-1]
            col_1st = df_out.columns[0]
            col_2nd = df_out.columns[1]

            df_out.loc[tail_line_number + 1, col_1st] = '表头'
            if header is not None:
                df_out[col_2nd] = df_out[col_2nd].astype(object)
                df_out.loc[tail_line_number + 1, col_2nd] = header.strip()
            else:
                df_out.loc[tail_line_number + 1, col_2nd] = 'None'
            df_out.loc[tail_line_number + 2, col_1st] = '表名'
            df_out.loc[tail_line_number + 2, col_2nd] = sheet_name.strip()
            df_out.loc[tail_line_number + 3, col_1st] = '文件名'
            df_out.loc[tail_line_number + 3, col_2nd] = file_name.name.strip()
            df_out.loc[tail_line_number + 4, col_1st] = ''

            total_df_list.append(df_out)
            return

    def personalized(df):

        # 检查列名是否重复
        if df.columns.duplicated().any():
            # 只对重复出现的列名添加重复次数的数字，而只出现一次的列名保持不变
            df.columns = [f'{col}_{df.columns.tolist()[:i].count(col)+1}'
                          if df.columns.tolist()[:i].count(col) > 0
                          else col for i, col in enumerate(df.columns)]

        for col in df.columns:

            for j in df[col]:
                # 如果列中元素在医生名单列表中，则该列为"住院医师"列
                if str(j) in lst_doctors:
                    df = df.copy()

                    personalized_df_list.append(df)

                    # 收集罚款数据
                    # 将所有医生列统一重命名为住院医师
                    df = df.rename(columns={col: "住院医师"})
                    list_col = df.columns

                    """保留包含在fine_list和reward_list列表中的列名和'住院医师'列"""
                    # 构建正则表达式，每个关键词前加上^表示行的开始，使用re.escape来转义关键词中可能存在的正则特殊字符
                    pattern = r'^(' + '|'.join(
                        re.escape(keyword) for keyword
                        in (fine_list + reward_list)) + ')'
                    # print(list_col)
                    additional_list = [col for col in list_col
                                       if isinstance(col, str)
                                       and re.search(pattern, col)
                                       and 'DRG结算金额' not in str(col)]
                    selected_columns_set = set(['住院医师'] + additional_list)

                    selected_columns = list(selected_columns_set)
                    if len(selected_columns) > 1:
                        # 填充空值
                        df["住院医师"] = df["住院医师"].ffill()
                        df_selected = df[selected_columns]
                        if len(df_selected) > 0:
                            df_selected = df_selected.copy()
                            for cl in [i for i in df_selected.columns
                                       if i != '住院医师']:
                                # 去除数字列中的汉字
                                df_selected[cl] = df_selected[cl].apply(
                                    lambda x: re.sub(r'\D', '', x) if
                                    isinstance(x, str) else x
                                    )
                                # 将列转换为数值类型，并处理非数字数据
                                df_selected[cl] = pd.to_numeric(
                                    df_selected[cl], errors='coerce')
                                # 对数字列中的数字取绝对值
                                df_selected[cl] = df_selected[cl].abs()

                                # 如果是奖励则将其值设置为负
                                if cl in reward_list:
                                    df_selected[cl] = df_selected[cl] * -1

                            # 当行中至少包含2个非空值时才保留该行，其他包含一个或更少非空值的行将被删除。
                            df_selected.dropna(axis=0, thresh=2, inplace=True)
                            # 获得的df收集到一个列表中，到最后concat输出
                            total_lst.append(df_selected)
                    break
        return

    def set_style_output(df_list, path_out):
        # 创建一个新的 Workbook 对象
        wb = Workbook()
        ws = wb.active

        for df in df_list:
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

        # 设置文字样式
        font_bold_red = Font(bold=True, color='FF0000')
        font_bold_blue = Font(bold=True, color='0000FF')
        font_bold_green = Font(bold=True, color='00FF00')
        # 设置黄绿色渐变背景样式
        green_yellow_fill = GradientFill(stop=("FF00FF00", "FFFFFF00"))

        # 遍历第一列单元格
        for row in ws.iter_rows(min_row=1, max_col=1, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    col = get_column_letter(cell.column)
                    ws.column_dimensions[col].width = 10
                    cell.font = font_bold_red
                    cell.fill = green_yellow_fill

        # 遍历所有单元格
        r = ws.max_row
        c = ws.max_column

        """分别定义罚款和奖励的 pattern"""
        # 以罚款列表中的元素开头
        pattern_fine = r'^(' + '|'.join(
            re.escape(keyword) for keyword in fine_list) + ')'
        # 存在奖励列表中的元素，注意：不限于开头
        pattern_reward = r'(' + '|'.join(
            re.escape(keyword) for keyword in reward_list) + ')'

        for row in range(1, r + 1):
            for column in range(1, c + 1):
                cell = ws.cell(row, column)
                cv = cell.value
                if not pd.isna(cv) and type(cv) is str:
                    if re.search(pattern_fine, cv):
                        cell.font = font_bold_red
                    if re.search(pattern_reward, cv):
                        cell.font = font_bold_green
                    if cv in lst_doctors:
                        cell.font = font_bold_blue
                    if cv == '表头':
                        ws.cell(row, column + 1).font = Font(bold=True)

        # 保存 Workbook 对象到 Excel 文件

        file_out = path_out.split('\\')[-1].split('.')[0]
        print(f"{file_out} ... ", end="")
        wb_save(wb, path_out)
        wb.close()

    def get_fine():

        # 确认 resules 文件夹存在
        Path(os.path.join(path, "results")).mkdir(parents=True, exist_ok=True)

        # 获取含有相关科室的df,收集到一个列表中
        path_f = Path(os.path.join(path, 'Fine'))
        file_list = [i for i in path_f.iterdir()
                     if (i.suffix in ['.xlsx', '.xls']
                         and '~' not in i.stem
                         )]
        for file_name in file_list:
            sheet_names = pd.ExcelFile(file_name).sheet_names
            for sheet_name in sheet_names:
                df = pd.read_excel(file_name, sheet_name=sheet_name)
                if not df.empty and 'SQLStatement' not in sheet_name:
                    # 去除df中所有单元格中的空格和工号
                    df = wipe(df)
                    header = None
                    # 使用 df.columns.str.contains() 方法来筛选列名
                    filtered_columns = df.columns[
                        ~df.columns.str.contains('Unnamed')]
                    # 统计不含有 "Unnamed" 的元素的数量
                    count = len(filtered_columns)
                    if count == 0:  # 该列为空
                        header = 'Empty1stRow HaveCol'
                        # print('The first row is empty.')
                        df_str = df.to_string(index=False)
                        result = any(item in df_str for item in lst_dep)
                        if not result:
                            header = 'Empty1stRow AddCol'
                            # print('There is no dep in the df.')
                            df = if_no_columns(df, file_name, sheet_name)
                            # 加入columns送入filter
                            filter(df, file_name, sheet_name, header)
                    if count > 1:  # 该列为columns，但不一定是需要的columns
                        # 查找columns内是否有科室列表lst_dep中的元素
                        found = [dep for dep in lst_dep if dep in df.columns]
                        if len(found) == 0:
                            header = None
                            # print('The first row is not columns.')
                            df_str = df.to_string(index=False)
                            result = any(item in df_str for item in lst_dep)
                            if not result:
                                header = '1stRowIsNotCol AddCol'
                                # print('There is no dep in the df.')
                                df = if_no_columns(df, file_name, sheet_name)
                                # 加入columns后送入filter
                                filter(df, file_name, sheet_name, header)
                        else:
                            header = '1stRowIsCol'
                            # 第一行是columns，直接送入filter
                            # print('The first row IS columns.')
                            filter(df, file_name, sheet_name, header)
                    if header is None:
                        # 到这时的df第一列为表头列，只获取第一非空单元格做为表头，其它表头忽略
                        first_non_Unnamed_index = numpy.where(
                            ~df.columns.str.contains('Unnamed'))[0][0]
                        header = df.columns[first_non_Unnamed_index]
                    # 将df切片输出df列表
                    split_dfs = df_spliter(df)

                    # 列表内df_slice逐个执行filter操作
                    for df_slice in split_dfs:
                        filter(df_slice, file_name, sheet_name, header)

        set_style_output(total_df_list,
                         os.path.join(path, r'results\全科汇总.xlsx'))

        set_style_output(personalized_df_list,
                         os.path.join(path, r'results\个人相关.xlsx'))

        if len(total_lst) > 0:
            # 合并 total_lst 中的所有 df
            concatenated_df = pd.concat(total_lst, ignore_index=True)
            out_df = concatenated_df.set_index('住院医师')
            # 有时有nan空列，删除之
            out_df = out_df.dropna(axis=1, how="all")
            print('个人明细 ... ', end="")
            save_xlsx(out_df, os.path.join(path, r'results\个人明细.xlsx'))

            df_gp = df_groups[["G", "D"]].dropna(subset="D").copy()
            summarized_df = summarize(concatenated_df, '罚款到组', df_gp)

        else:
            summarized_df = df_null("罚款到组")

        save_xlsx(summarized_df, os.path.join(path, r'results\罚款到组.xlsx'))
        return summarized_df

    print()
    print("获取罚款相关信息……")
    return get_fine()


# 汇总三项+二项收入数据
class revenue_summary:

    def __init__(self, path):

        self.path = path
        Path(os.path.join(path, "results")).mkdir(parents=True, exist_ok=True)
        
        df_groups = get_df_groups(path)[["G", "D"]].dropna(subset="D")

        # etc 组的设定是为了区别其它组，对奖金分配没有影响
        staff_lst = df_groups.D.unique().tolist()
        self.df_groups = df_groups[df_groups["G"].isin(staff_lst)][["G", "D"]]

        # 获取脚本所在的目录
        script_dir = get_abspath()
        # 获取sumlist文件的绝对路径
        sumlistPath = os.path.join(script_dir, r'lists\sumlist.xlsx')
        check_if_exists(sumlistPath)
        self.df_sumlist = pd.read_excel(sumlistPath).fillna(1)

    # '吉林省人民医院病人记帐费用汇总清单'
    def detailed_list(self, path_src):

        df_src = pd.read_excel(path_src)

        msg = '住院清单'
        df_src.columns = df_src.loc[1]
        df = df_src.drop([0, 1])

        # 精简需要的列
        df = df[['住院医师'] + [i for i in df.columns
                            if i in self.df_sumlist['G'].values]]
        # 收入按比例折算
        for i in df.columns:
            for j, v in enumerate(self.df_sumlist['G'].values):
                if i == v:
                    df[i] *= self.df_sumlist['P'][j]
        print()
        df_out = summarize(df, msg, self.df_groups)
        return df_out

    # '吉林省人民医院科室医生收入统计表(开单科室)'
    def outpatient_revenue(self, path_src):

        df_src = pd.read_excel(path_src)
        msg = '门诊检查'
        df = df_src.T
        df.columns = df.iloc[0]
        df.reset_index(inplace=True)
        df = df.drop([0])
        df = df[['收入项目', '检查收入', '治疗收入']]
        # df = df.rename(columns={df.columns[0]: '住院医师'})
        df.columns = ['住院医师', '检查收入', '治疗收入']

        df_out = summarize(df, msg, self.df_groups)
        return df_out

    # '吉林省人民医院科室手术统计分析表'
    def inpatient_surgery(self, path_src):

        df_src = pd.read_excel(path_src)
        df_src.columns = df_src.iloc[1]

        """获取非机器人数据（普通手术）"""
        msg_norm = '手术分析'
        df_norm = df_src[~df_src['项目'].str.contains('机器人', na=False)]
        df_norm = df_norm.drop([0, 1])
        df_norm = df_norm[['住院医师', '金额']]
        df_norm = summarize(df_norm, msg_norm, self.df_groups)

        """获取机器人数据"""
        df_robot = df_src[df_src['项目'].str.contains('机器人', na=False)]

        # 获取机器人收入
        msg_rob = '机器人'
        df_rob = df_robot[['住院医师', '金额']]
        df_rob = summarize(df_rob, msg_rob, self.df_groups)

        # 获取机器人病例
        # 先去除“住院医师”列中单元格的所有空格和工号信息
        df_robdata = pd.merge(
            self.df_groups, wipe(df_robot),
            left_on="D", right_on='住院医师',
            how="inner"
            )
        df_robdata.drop("D", axis=1, inplace=True)
        df_robdata.set_index(["G", "住院医师"], inplace=True)
        print("rob.xlsx ... ", end="")
        save_xlsx(df_robdata, os.path.join(self.path, r"results\rob.xlsx"))

        return df_norm, df_rob

    # 门诊手术
    def outpatient_treatments(self, path_src):

        msg = '门诊手术'

        if not Path(path_src).exists():
            return df_null(msg)

        df_src = pd.read_excel(path_src)

        if "绩效" not in df_src.to_string(index=False):
            return df_null(msg)

        y, x = df_src.shape

        # 定位医生列
        for i in range(x):
            set_outpatient_doctor = set(
                item.replace(" ", "")  # 只有当item是字符串时，才调用replace方法
                for item in df_src.iloc[:, i].dropna()  # 排除NaN值
                if isinstance(item, str)  # 确保item是字符串
            )
            # 判断两个集合是否有交集
            if set_outpatient_doctor & set(self.df_groups.D):
                col_doctor = i
                break

        # 定位绩效列
        for i, j in enumerate(df_src.columns):
            if "绩效" in str(j):
                df = df_src.iloc[:, [col_doctor, i]]
                break
            else:
                for j in range(y):
                    for k in range(x):
                        cell_value = df_src.iloc[j, k]
                        if "绩效" in str(cell_value):
                            df = df_src.iloc[:, [col_doctor, k]]
                            break

        df.columns = ['住院医师', '医生绩效30%']

        df_out = summarize(df, msg, self.df_groups)
        return df_out

    # 挂号会诊
    def Rigistration_Consultation(self, path_src):

        msg_reg = '平日挂号'
        msg_con = '会诊'

        if not Path(path_src).exists():
            return df_null(msg_reg), df_null(msg_con)

        df_src = pd.read_excel(path_src, index_col="住院医师").fillna(0)
        df_reg = df_src[["Registration"]]
        df_out_reg = summarize(df_reg, msg_reg, self.df_groups)

        df_con = df_src[["Consultation"]]
        df_con_reg = summarize(df_con, msg_con, self.df_groups)

        return df_out_reg, df_con_reg


# 汇总三项+二项收入数据-输出
def get_all(path):

    path_1 = r'ZLHIS\吉林省人民医院病人记帐费用汇总清单_ForUse.xlsx'
    if not Path(os.path.join(path, path_1)).exists():
        patient_affiliation_check(path)
    path_2 = r'ZLHIS\吉林省人民医院科室医生收入统计表(开单科室).xlsx'
    path_3_4 = r'ZLHIS\吉林省人民医院科室手术统计分析表.xlsx'
    path_5 = r'ZLHIS\门诊手术.xlsx'
    path_6_7 = r'ZLHIS\挂号会诊.xlsx'

    calc = revenue_summary(path)

    return \
        calc.detailed_list(os.path.join(path, path_1)), \
        calc.outpatient_revenue(os.path.join(path, path_2)), \
        * calc.inpatient_surgery(os.path.join(path, path_3_4)), \
        calc.outpatient_treatments(os.path.join(path, path_5)), \
        * calc.Rigistration_Consultation(os.path.join(path, path_6_7))


# 汇总三项+二项收入数据-保存并打印
def output_revenue_data(path):

    dfs = get_all(path)
    df_out = pd.concat(dfs, axis=1)
    # 设置格式
    df_out_with_style = set_gradient_style(df_out)
    path_out = os.path.join(path, r'results\0_merged_ratio.xlsx')
    print("0_merged_ratio.xlsx original data", end="")
    save_xlsx(df_out_with_style, path_out)
    print()
    print(tabulate(df_out, headers='keys', tablefmt='pipe'))


# 整合所有收入及奖罚数据，保存到两个文件中 0_merged_ratio、income_penalty
def Data_sum(path):

    path_income_penalty = os.path.join(path, "income_penalty.xlsx")
    check_if_exists(path_income_penalty)

    if not os.path.exists(path_income_penalty):
        # print()
        print("\nFile (income_penalty.xlsx) path error.\nPlease try again.\n")
        return

    df_income_penalty = pd.read_excel(path_income_penalty,
                                      index_col="收入和处罚").fillna(0)

    df_nurses_info, df_doctors_info = staff_info(path)
    df_在勤 = df_doctors_info[(~df_doctors_info.S.isin(["返聘", "不在岗"])) \
        & (df_doctors_info.PW != 0) & (df_doctors_info.PG != 0)]
    出勤系数 = df_在勤.PW.sum()

    """填加spacer"""
    df_income_penalty.loc["*" * 7, "金额"] = 0

    """填入相对固定数据"""
    # 获取住院清单中的日期范围及当月天数
    dates, days_in_month = get_Date_range(path)
    # 计算夜班费、医生出勤及中层金额并填入文档
    df_income_penalty.loc["保洁", "金额"] = 440
    df_income_penalty.loc["中层", "金额"] = 240 + 240/2
    df_income_penalty.loc["夜班费", "金额"] = days_in_month * 100
    df_income_penalty.loc["医生出勤", "金额"] = 出勤系数 * 200

    医生总值班费 = extract_namelist_from_word(path)
    df_income_penalty.loc["总值", "金额"] = 医生总值班费

    """填加spacer"""
    df_income_penalty.loc["*" * 8, "金额"] = 0

    # 获取收入数据
    # [sumarizing, income, surgery, robot, out_surg]
    dfs_output = get_all(path)

    """填入PRP金额"""
    df_income_penalty.loc["PRP", "金额"] = getPRP(path)
    print("PRP .... Done.")

    """填入机器人金额 0.28 """
    df_income_penalty.loc["机器人", "金额"] = int(dfs_output[3]
                                             .loc["总计", "合计"] * 0.28)

    """填入门诊手术金额"""
    df_income_penalty.loc["门诊手术", "金额"] = int(dfs_output[4]
                                              .loc["总计", "合计"])

    # 分组罚款，科室处罚和奖励
    fine_of_groups = query_fine_data(path)

    """填入罚款数据"""
    fine = fine_of_groups.loc["总计", "合计"]
    df_income_penalty.loc["罚款到组", "金额"] = fine

    total_fine = df_income_penalty.loc["处罚", "金额"]
    df_income_penalty.loc["科室处罚", "金额"] = total_fine - fine

    df_income_penalty["金额"] = df_income_penalty["金额"].apply(int)

    print("income_penalty.xlsx ... ", end="")
    # 保存总体收入数据
    save_xlsx(df_income_penalty, path_income_penalty)

    """将罚款项目加入dfs列表并横向合并"""
    df_merged = pd.concat([*dfs_output, fine_of_groups], axis=1)

    # 设置格式
    df_merged = set_gradient_style(df_merged)

    # 保存收入比例数据
    path_merged = os.path.join(path, r"results\0_merged_ratio.xlsx")
    print("0_merged_ratio.xlsx ... ", end="")
    save_xlsx(df_merged, path_merged)

    # 打印计算的月份
    print(f'bonus_calculation date: {dates}')


# 最终奖金计算
def bonus_calculation(path):

    # ---------------------------------------------------------------------------------------------------------------

    """读入收入数据及分配比例"""

    df_nurses_info, df_doctors_info = staff_info(path)

    """全科收入数据"""
    path_income_penalty = os.path.join(path, r"income_penalty.xlsx")
    check_if_exists(path_income_penalty)
    df_income_penalty = pd.read_excel(path_income_penalty,
                                      index_col="收入和处罚")

    """各组收入比例数据"""
    path_merged_ratio = os.path.join(path, r"results\0_merged_ratio.xlsx")
    check_if_exists(path_merged_ratio)
    staff_lst = df_doctors_info.D.unique().tolist()
    df_merged_ratio = pd.read_excel(path_merged_ratio, index_col="G")\
        .loc[lambda df: df.index.isin(staff_lst)]


    """获取各种工作状态分组数据"""

    # region

    df_在岗, gb_在岗, df_在勤, gb_在勤, df_夜班, gb_夜班 = info_df_gb(df_doctors_info)

    print(f'\nGroups: {df_doctors_info.groupby("G").agg(list).D}\n')

    print("If there are any member adjustments,",
          "please remember to modify the 'groups.xlsx' file.\n")
    # ---------------------------------------------------------------------------------------------------------------

    # 对各类人数进行判断

    医生总数 = len(df_doctors_info.D)

    # 求出各种情况的人数
    counter = Counter(df_doctors_info.S)

    返聘人数 = counter["返聘"]
    在职人数 = 医生总数 - 返聘人数

    在勤系数 = gb_在勤.PW.sum()
    部分勤系数 = gb_在勤.loc["partial", "PW"]\
        if "partial" in gb_在勤.index else 0

    部分奖比例 = 部分勤系数 / 在勤系数

    不在岗数 = counter["不在岗"]
    均奖人数 = counter["均奖"]

    在岗系数 = gb_在岗.PW.sum()
    均奖人比例 = 均奖人数 / 在岗系数
    临床人比例 = 1 - 均奖人比例 - 部分奖比例

    nurses = set(df_nurses_info.N)
    全体护士人数 = len(nurses)

    副主任系数 = 0.6
    护长系数 = 0.5

    医生系数 = 在职人数 + 副主任系数
    护士系数 = 全体护士人数 * 1.2 + 护长系数

    out_items = ["医生总数", "返聘人数", "在职人数", "在勤系数", "不在岗数", "均奖人数", "医生系数"]

    for i in out_items:
        v = eval(i)
        if v > 0:
            print(f"{i}: {v}")
    print()
    print(f"护士总数: {全体护士人数}\n护士系数: {护士系数}\n")

    总体系数 = round((医生系数 + 护士系数), 1)

    with open(os.path.join(path, '总值名单.txt'), 'r', encoding='UTF-8') as f:
        content = f.read()
        创二总值 = content.split(',')
        # 获取医生总值名单列表
        医生总值 = [item for item in 创二总值 if item not in nurses]

    # endregion

    # ===============================================================================================================

    """汇总结果"""
    # 汇总(组)
    df_preliminary = pd.DataFrame()
    # 汇总（人）
    df_result = pd.DataFrame()
    # 护士
    df_nurses = pd.DataFrame()

    # ***************************************************************************************************************

    """汇总(组)"""

    # region

    """按比例分配"""
    # 定义一空字典，然后向其中添加收入（key）和参照比例（value）
    dict_data_prorated = dict()
    # 1、科室处罚
    科室处罚 = -df_income_penalty.loc["科室处罚", "金额"]
    dict_data_prorated["科室处罚"] = "住院清单"    # 填入字典中 +++++++++++++++++++++++++++++++++ 1

    # 2、罚款到组
    罚款到组 = -df_income_penalty.loc["处罚", "金额"] - 科室处罚
    dict_data_prorated["罚款到组"] = "罚款到组"    # 填入字典中 +++++++++++++++++++++++++++++++++ 2

    # 3、门诊检查
    门诊检查 = int(df_income_penalty.loc["门诊收入之检查", "金额"] * 0.2)
    dict_data_prorated["门诊检查"] = "门诊检查"    # 填入字典中 +++++++++++++++++++++++++++++++++ 3

    # 4、门诊手术
    门诊手术 = df_income_penalty.loc["门诊手术", "金额"]
    dict_data_prorated["门诊手术"] = "门诊手术"    # 填入字典中 +++++++++++++++++++++++++++++++++ 4

    # 5、手术分析
    手术分析 = df_income_penalty.loc["手术(含门诊)", "金额"] - 门诊手术
    dict_data_prorated["手术分析"] = "手术分析"    # 填入字典中 +++++++++++++++++++++++++++++++++ 5

    机器人28 = df_income_penalty.loc["机器人", "金额"]
    # 6、机器人20
    机器人20 = 机器人28 / 28 * 20  # 按比例
    dict_data_prorated["机器人20"] = "机器人"      # 填入字典中 +++++++++++++++++++++++++++++++++ 6
    机器人8 = int(机器人28 / 28 * 8)  # 给护士
    # 机器人另8%都给护士

    # 7. 平日挂号
    # 8. 会诊
    平日挂号 = df_income_penalty.loc["平日挂号费提成", "金额"]
    dict_data_prorated["平日挂号"] = "平日挂号"     # 填入字典中 +++++++++++++++++++++++++++++++++ 7
    会诊 = df_income_penalty.loc["会诊费提成", "金额"]
    dict_data_prorated["会诊"] = "会诊"            # 填入字典中 +++++++++++++++++++++++++++++++++ 8

    # 9. 医生绩效 --------------------------------------------------------------
    绩效奖金 = df_income_penalty.loc["绩效奖金", "金额"]
    其它 = df_income_penalty.loc["其它", "金额"]
    奖励 = df_income_penalty.loc["奖励", "金额"]
    保洁 = df_income_penalty.loc["保洁", "金额"]
    PRP = df_income_penalty.loc["PRP", "金额"]

    if 门诊手术 == 0:
        print("\n门诊手术金额为 0 ，请注意核查。")

    # 绩效构成如下：
    # (计奖净收入-支出)x门诊占比x30%]+[(计奖净收入-支出)×(1-门诊占比)x25%]+门诊检查费+节假日挂号(80%)+平日挂号(50%)+机器人开机(28%)+会诊费(50%)

    # 按劳分配绩效如下 (其中 门诊检查 = 门诊收入之检查 * 20%)：
    # 绩效 = 绩效奖金 + 其它 + 奖励 - 保洁 - 门诊检查 - 机器人28 - 平日挂号 - 会诊
    # 以下用循环方法合并字符串
    performance_contents_addition = ["绩效奖金", "其它", "奖励"]  # 相加
    performance_contents_subtract = ["保洁", "门诊检查", "机器人28", "平日挂号", "会诊"]   # 相减

    # 求和，合并字符串
    绩效 = 0
    string_performance_contents = ""
    string_performance_contents_revernue = ""

    for i, j in enumerate(performance_contents_addition):
        revenue = eval(j)
        绩效 += revenue
        if i == 0:
            string_performance_contents += j
            string_performance_contents_revernue += f'{revenue}'
        else:
            string_performance_contents += ' + ' + j
            string_performance_contents_revernue += f' + {revenue}'

    for i, j in enumerate(performance_contents_subtract):
        revenue = eval(j)
        绩效 -= revenue
        string_performance_contents += ' - ' + j
        string_performance_contents_revernue += f' - {revenue}'

    人均绩效 = round(绩效 / 总体系数, 1)

    绩效信息 = "绩效全科人均" + "\n"\
          f"({string_performance_contents}) / {总体系数}" + "\n"\
          f"({string_performance_contents_revernue}) / {总体系数}" + "\n"\
          f"(人均绩效) = {人均绩效}" + "\n"

    print(绩效信息)

    医生绩效 = (绩效 - PRP) / 总体系数 * 在职人数  # 其中的"医生PRP"之后平分

    dict_data_prorated["医生绩效"] = "住院清单"     # 填入字典中 +++++++++++++++++++++++++++++++++ 9

    护士绩效 = int(人均绩效 * 护士系数)

    print(f'Total performance (去除"机器人28"和"PRP")：{医生绩效.round()}\n')

    # endregion

    # -----------------------------------------------------------------------------------------

    """ 字典 dict_data_prorated 填充完成 """

    """ 根据dict_data_prorated计算各组（除etc外）的收入，填入df_preliminary中 """

    # region

    # 按比例分配数据 dict 填入 df_preliminary
    for k, v in dict_data_prorated.items():
        amount = eval(k) * df_merged_ratio[v]
        df_preliminary[k] = amount * 临床人比例

    allocation_items = {}
    """计算均奖人收入，如果有则合并"""
    if 均奖人数 > 0:
        # 计算 均奖 组的收入，keys 为收入或罚款项目金额名称，allocation_items to df_avg_group
        for k in dict_data_prorated.keys():
            allocation_items[k] = eval(k) * 均奖人比例

    # 均奖组数据 dict 转 df
    df_avg_group = pd.DataFrame(allocation_items, index=["avg"])

    partial_items = {}
    """计算部分奖人收入，如果有则合并"""
    if 部分勤系数 > 0:
        for k in dict_data_prorated.keys():
            partial_items[k] = eval(k) * 部分奖比例

    # 部分勤组数据 dict 转 df
    df_part_group = pd.DataFrame(partial_items, index=["partial"])

    # 合并三个 df
    df_preliminary = pd.concat([df_preliminary, df_avg_group, df_part_group])

    # -----------------------------------------------------------------------------------------

    """所有医生平均分配项目"""

    # 定义一个空列表 list_average_items ，存放平均分配项目，传参，后面汇总时还要用到
    list_average_items = list()
    # 1. PRP 医生护士绩效后平均
    医生PRP = PRP / 总体系数 * 在职人数
    list_average_items.append("医生PRP")       # 加入列表中 =================== 1

    # 2. LK中层 医生占一半，为120
    LK中层 = 120
    list_average_items.append("LK中层")        # 加入列表中 =================== 2

    """ 列表 list_average_items 填充完成 """

    ###########################################################################################

    """将数据填入到 df_preliminary 中"""
    for i in list_average_items:
        df_preliminary[i] = eval(i) / 在岗系数 * gb_在岗.PW

    # -----------------------------------------------------------------------------------------

    """参与者平均分配"""
    # 1. 值夜班者平均
    夜班费 = df_income_penalty.loc["夜班费", "金额"]
    # 夜班费每组平均
    df_preliminary.loc[gb_夜班.index, "夜班费"] = 夜班费 / len(gb_夜班)

    # 2. 在职者平均
    # 全勤奖
    医生出勤 = df_income_penalty.loc["医生出勤", "金额"]  # 200 * 在勤系数
    df_preliminary["医生出勤"] = 医生出勤 / 在勤系数 * gb_在勤.PW

    """项目存入一个列表后面汇总时要用到"""
    dict_avg_item_dfs = {"夜班费": gb_夜班, "医生出勤": gb_在勤}

    # -----------------------------------------------------------------------------------------

    """特殊津贴"""

    # 通过 lst_sb 获取其所在组名idx(s)， col 为 list_sb 所在列名
    def group_idx(gb_df, lst_sb, col):
        # 返回值也是一个 list
        return gb_df[gb_df[col].apply(
            lambda x: any(name in x for name in lst_sb))].index

    # 1. 主任
    # 李光淳所在组的索引 lgc_group
    lgc_group = group_idx(gb_在岗, ["李光淳"], "D_list")
    光淳均奖 = 人均绩效 * 0.6
    df_preliminary.loc[lgc_group, "光淳均奖"] = 光淳均奖
    光淳中层 = 240
    df_preliminary.loc[lgc_group, "光淳中层"] = 光淳中层

    # 2. 院聘
    def boolean_zsz():
        return "赵栓柱" in staff_lst

    if boolean_zsz():
        柱子聘金 = 5000
        zsz_group = group_idx(gb_在岗, ["赵栓柱"], "D_list")
        df_preliminary.loc[zsz_group, "柱子聘金"] = 柱子聘金

    # 3. 总值
    # 总值所在组的索引 lst_idx_zz
    总值 = df_income_penalty.loc["总值", "金额"]\
        if "总值" in df_income_penalty.index else 0  # 如果总值费用未记入，则暂记为 0

    if "" not in 医生总值:
        lst_idx_zz = group_idx(gb_在岗, 医生总值, "D_list")
        # 总值 / len(lst_idx_zz) = 50
        df_preliminary.loc[lst_idx_zz, "总值"] = 50
    else:
        df_preliminary.loc["Total", "总值"] = 0

    """将 df_preliminary 按行列分别求和"""
    df_preliminary.loc["Total"] = df_preliminary.sum()
    df_preliminary.loc[:, "奖金"] = df_preliminary.sum(axis=1)

    df_preliminary = df_preliminary.round()

    # endregion

    # &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&

    """汇总（人） 根据 df_preliminary 完成 df_result """

    # region

    # 填充按比例分配项目，k 为收入项目，组成员平均(组内比例)

    for k in dict_data_prorated.keys():
        if eval(k) != 0:
            for leader in gb_在岗.index:
                amount = df_preliminary.loc[leader, k]  # 获取金额
                members = gb_在岗.loc[leader, "D_list"]  # 组成员
                for d in members:
                    # 个体系数
                    p = df_在岗.query(f"D == '{d}'").PG.values[0]
                    # 组总系数
                    c = gb_在岗.loc[leader, "PG"]
                    df_result.loc[d, k] = amount / c * p

    # 所有医生平均分配项目，i 为金额名称，包括返聘，排除不在岗者
    # 组内依然按比例及系数比分配

    # 定义一个函数，通过D列的值获取PG列的值
    def pg_values(df, d):
        # 过滤出D列等于d_value的行
        filtered_df = df[df['D'] == d]
        # 获取对应的PG值
        pg_value = filtered_df['PG'].tolist()[0]
        return pg_value

    list_在岗 = df_在岗.D.tolist()
    for i in list_average_items:
        for d in list_在岗:

            leader = df_doctors_info.query(f"D == '{d}'").G.values[0]
            科总系数 = df_在岗.PW.sum()
            if leader == "avg":
                df_result.loc[d, i] = eval(i) / 科总系数

            else:
                个人科内系数 = df_在岗.query(f"D == '{d}'").PW.values[0]
                个人组内系数 = pg_values(df_在岗, d)
                组内人数 = gb_在岗.loc[leader, "D_count"]
                组总系数 = gb_在岗.loc[leader, "PG"]

                # if 组总系数 == 组内人数:
                #     df_result.loc[d, i] = eval(i) / 科总系数 * 个人科内系数
                # else:
                #     df_result.loc[d, i] = eval(i) / 科总系数 * 组内人数 * (个人组内系数 / 组总系数)

                # 上面的if-else 判断可以合并成下面一句
                df_result.loc[d, i] = eval(i) / 科总系数 * 个人科内系数 * \
                    组内人数 * (个人组内系数 / 组总系数)

    # 参与者平均分配项目
    # k, v 分别为相对应的收入项目（夜班费和医生出勤）及分组信息

    for i in staff_lst:  # 先填加所有医生到 df_result 中，否则遇均奖人会报错
        if i not in df_result.index:
            df_result.loc[i] = pd.NA

    for k, v in dict_avg_item_dfs.items():

        list_d = sum(v['D_list'].tolist(), [])
        for d in list_d:
            leader = group_idx(v, [d], "D_list").values[0]
            length = len(v.loc[leader, "D_list"])
            amount = df_preliminary.loc[leader, k] / length
            df_result.loc[d, k] = amount

    # 特殊津贴
    df_result.loc["李光淳", "光淳均奖"] = 光淳均奖
    df_result.loc["李光淳", "光淳中层"] = 光淳中层

    if boolean_zsz():
        df_result.loc["赵栓柱", "柱子聘金"] = 柱子聘金

    if "" not in 医生总值:
        for i in 医生总值:
            df_result.loc[i, "总值"] = 50

    """将 df_result 按行列分别求和"""
    df_result.loc["Total"] = df_result.sum()
    df_result["奖金"] = df_result.sum(axis=1)

    # 去除df中全为0的列、取整
    df_result = (df_result.loc[:, (df_result != 0).any(axis=0)]).round()
    医生收入 = df_result.loc["Total", "奖金"]

    path_output = os.path.join(path, f"result_{os.path.basename(path)}.xlsx")
    performance_data_output = os.path.join(path, f"performance_{os.path.basename(path)}.txt")

    # 输出绩效信息到文本文件
    with open(performance_data_output, "w", encoding="utf-8") as f:
        f.write(绩效信息 + "\n")
    os.system(f'start notepad "{performance_data_output}"')

    print(f"请核对总值名单及金额：\n创二总值 {创二总值}，\n医生总值 {医生总值}，\n医生总值班费 {总值} 元。\n")

    # -----------------------------------------------------------------------------------------

    # 医生加护士与应发金额比较
    护士全勤 = df_income_penalty.loc["全勤奖", "金额"] - 医生出勤
    护士夜班 = 夜班费
    护长总值 = 50 if "汪珠" in 创二总值 else 0
    护长中层 = 240

    # 护士收入构成
    revenue_nurses = ["护士绩效", "机器人8", "LK中层", "护士全勤", "护士夜班", "护长总值", "护长中层", "保洁"]

    # 求和并将列表内容转换成字符串输出打印
    护士收入 = 0
    string_nurses = ""
    string_revenue_nurses = ""
    for i, j in enumerate(revenue_nurses):
        revenue = eval(j)
        df_nurses.loc[j, "护士收入"] = revenue
        护士收入 += revenue
        if i == 0:
            string_nurses += j
            string_revenue_nurses += f'{revenue}'
        else:
            string_nurses += ' + ' + j
            string_revenue_nurses += f' + {revenue}'

    df_nurses.loc["Total", "护士收入"] = 护士收入
    df_nurses.loc["Total", "医生收入"] = 医生收入
    total = 护士收入 + 医生收入
    df_nurses.loc["Total", "Total"] = total
    应发金额 = df_income_penalty.loc["应发金额", "金额"]
    df_nurses.loc["Total", "应发金额"] = 应发金额
    差值 = 应发金额 - total
    df_nurses.loc["Total", "差值"] = 差值

    print(f"护士收入 = {string_nurses}")
    print(f"{护士收入} = {string_revenue_nurses}\n")

    print(f'应发金额(Total) = 医生收入 + 护士收入')
    print(f'{int(total)} = {int(医生收入)} + {int(护士收入)}')
    print("应发金额：", f"{应发金额}, 差值: {差值}\n", sep="\n")

    # endregion

    # -----------------------------------------------------------------------------------------

    """输出excel"""

    # region

    while True:
        try:
            with pd.ExcelWriter(path_output) as writer:
                df_result.to_excel(writer, sheet_name="result")
                df_preliminary.to_excel(writer, sheet_name="preliminary")
                df_nurses.to_excel(writer, sheet_name="nurses")
                string_printout = "The result has been saved to the path"
                print(f'{string_printout} " {path_output} "')

                # 使用 os.system 打开 Excel 文件
                os.system(f'start excel "{path_output}"')

                break  # If no exception, break the loop

        except PermissionError:
            print("Error(res): The file is currently in use by another program.",
                    "Please close the file and try again.")
            input()

        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            break  # Break on unexpected errors

    # endregion

    # -----------------------------------------------------------------------------------------

    return df_preliminary, df_result


if __name__ == "__main__":

    try:
        ctrl_panel()
    except KeyboardInterrupt:
        print("\nThe program has been interrupted by the user.")

    # pathFile = get_abspath("path.txt")
    # check_if_exists(pathFile)
    # with open(pathFile, "r", encoding="UTF8") as f:
    #     path = f.read().strip()
    # output_revenue_data(path)

    # path_6 = r'ZLHIS\挂号会诊.xlsx'
    # calc = revenue_summary(path)
    # df_6, df_7 = calc.Rigistration_Consultation(os.path.join(path, path_6_7))